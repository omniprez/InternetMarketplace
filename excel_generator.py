import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def format_workbook(workbook):
    """
    Apply formatting to the workbook
    """
    # Get the active sheet
    sheet = workbook.active
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
    
    # Apply header style to the first row
    for cell in sheet[1]:
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill
    
    # Apply border to all cells with data
    data_rows = sheet.max_row
    data_cols = sheet.max_column
    
    for row in range(2, data_rows + 1):
        for col in range(1, data_cols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.font = normal_font
            
            # Align numbers to the right
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
    
    # Auto-adjust column width
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    return workbook

def create_excel_file(invoice_data, output_path, batch_mode=False):
    """
    Generate an Excel file from the extracted invoice data
    """
    logger.info(f"Creating Excel file at: {output_path}")
    
    try:
        if batch_mode:
            # Handle batch processing of multiple invoices
            create_batch_excel(invoice_data, output_path)
        else:
            # Handle single invoice
            create_single_invoice_excel(invoice_data, output_path)
            
        logger.info(f"Excel file created successfully at: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        raise

def create_single_invoice_excel(invoice_data, output_path):
    """
    Generate Excel file for a single invoice
    """
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice Data"
    
    # Add header information
    sheet['A1'] = "Invoice Information"
    sheet.merge_cells('A1:B1')
    
    header_data = invoice_data.get('header', {})
    row = 2
    
    sheet.cell(row=row, column=1).value = "Invoice Number"
    sheet.cell(row=row, column=2).value = header_data.get('invoice_number', '')
    row += 1
    
    sheet.cell(row=row, column=1).value = "Date"
    sheet.cell(row=row, column=2).value = header_data.get('date', '')
    row += 1
    
    sheet.cell(row=row, column=1).value = "Customer Reference"
    sheet.cell(row=row, column=2).value = header_data.get('customer_ref', '')
    row += 1
    
    sheet.cell(row=row, column=1).value = "Customer Name"
    sheet.cell(row=row, column=2).value = header_data.get('customer_name', '')
    row += 1
    
    sheet.cell(row=row, column=1).value = "VAT Registration Number"
    sheet.cell(row=row, column=2).value = header_data.get('vat_reg_no', '')
    row += 1
    
    sheet.cell(row=row, column=1).value = "Business Registration Number"
    sheet.cell(row=row, column=2).value = header_data.get('business_reg_no', '')
    row += 2
    
    # Add line items
    line_items = invoice_data.get('line_items', [])
    if line_items:
        sheet.cell(row=row, column=1).value = "Line Items"
        sheet.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Add line item headers
        headers = ["Product Code", "Quantity", "Description", "Unit Price", "Discount", "Total"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=row, column=col).value = header
        row += 1
        
        # Add line item data
        for item in line_items:
            sheet.cell(row=row, column=1).value = item.get('product_code', '')
            sheet.cell(row=row, column=2).value = item.get('quantity', '')
            sheet.cell(row=row, column=3).value = item.get('description', '')
            sheet.cell(row=row, column=4).value = item.get('unit_price', '')
            sheet.cell(row=row, column=5).value = item.get('discount', '')
            sheet.cell(row=row, column=6).value = item.get('total', '')
            row += 1
    
    row += 1
    
    # Add totals
    totals = invoice_data.get('totals', {})
    sheet.cell(row=row, column=5).value = "Subtotal"
    sheet.cell(row=row, column=6).value = totals.get('subtotal', '')
    row += 1
    
    sheet.cell(row=row, column=5).value = "VAT"
    sheet.cell(row=row, column=6).value = totals.get('vat', '')
    row += 1
    
    sheet.cell(row=row, column=5).value = "Grand Total"
    sheet.cell(row=row, column=6).value = totals.get('grand_total', '')
    
    # Apply formatting
    workbook = format_workbook(workbook)
    
    # Save the workbook
    workbook.save(output_path)

def create_batch_excel(invoice_data_list, output_path):
    """
    Generate Excel file for multiple invoices
    """
    # Create a new workbook
    workbook = Workbook()
    
    # Create summary sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    # Set up summary headers
    summary_headers = ["Invoice Number", "Date", "Customer", "Subtotal", "VAT", "Grand Total"]
    for col, header in enumerate(summary_headers, 1):
        summary_sheet.cell(row=1, column=col).value = header
    
    summary_row = 2
    
    # Process each invoice
    for i, invoice_data in enumerate(invoice_data_list):
        # Add invoice to summary
        header = invoice_data.get('header', {})
        totals = invoice_data.get('totals', {})
        
        summary_sheet.cell(row=summary_row, column=1).value = header.get('invoice_number', f'Invoice {i+1}')
        summary_sheet.cell(row=summary_row, column=2).value = header.get('date', '')
        summary_sheet.cell(row=summary_row, column=3).value = header.get('customer_name', '')
        summary_sheet.cell(row=summary_row, column=4).value = totals.get('subtotal', '')
        summary_sheet.cell(row=summary_row, column=5).value = totals.get('vat', '')
        summary_sheet.cell(row=summary_row, column=6).value = totals.get('grand_total', '')
        
        summary_row += 1
        
        # Create detail sheet for this invoice
        invoice_number = header.get('invoice_number', f'Invoice {i+1}')
        detail_sheet = workbook.create_sheet(f"INV-{i+1}")
        
        # Add header information
        detail_sheet['A1'] = f"Invoice: {invoice_number}"
        detail_sheet.merge_cells('A1:B1')
        
        row = 2
        detail_sheet.cell(row=row, column=1).value = "Invoice Number"
        detail_sheet.cell(row=row, column=2).value = header.get('invoice_number', '')
        row += 1
        
        detail_sheet.cell(row=row, column=1).value = "Date"
        detail_sheet.cell(row=row, column=2).value = header.get('date', '')
        row += 1
        
        detail_sheet.cell(row=row, column=1).value = "Customer Reference"
        detail_sheet.cell(row=row, column=2).value = header.get('customer_ref', '')
        row += 1
        
        detail_sheet.cell(row=row, column=1).value = "Customer Name"
        detail_sheet.cell(row=row, column=2).value = header.get('customer_name', '')
        row += 2
        
        # Add line items
        line_items = invoice_data.get('line_items', [])
        if line_items:
            detail_sheet.cell(row=row, column=1).value = "Line Items"
            detail_sheet.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Add line item headers
            headers = ["Product Code", "Quantity", "Description", "Unit Price", "Discount", "Total"]
            for col, header in enumerate(headers, 1):
                detail_sheet.cell(row=row, column=col).value = header
            row += 1
            
            # Add line item data
            for item in line_items:
                detail_sheet.cell(row=row, column=1).value = item.get('product_code', '')
                detail_sheet.cell(row=row, column=2).value = item.get('quantity', '')
                detail_sheet.cell(row=row, column=3).value = item.get('description', '')
                detail_sheet.cell(row=row, column=4).value = item.get('unit_price', '')
                detail_sheet.cell(row=row, column=5).value = item.get('discount', '')
                detail_sheet.cell(row=row, column=6).value = item.get('total', '')
                row += 1
        
        row += 1
        
        # Add totals
        detail_sheet.cell(row=row, column=5).value = "Subtotal"
        detail_sheet.cell(row=row, column=6).value = totals.get('subtotal', '')
        row += 1
        
        detail_sheet.cell(row=row, column=5).value = "VAT"
        detail_sheet.cell(row=row, column=6).value = totals.get('vat', '')
        row += 1
        
        detail_sheet.cell(row=row, column=5).value = "Grand Total"
        detail_sheet.cell(row=row, column=6).value = totals.get('grand_total', '')
    
    # Apply formatting to all sheets
    for sheet in workbook:
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
        
        # Apply header style to the first row
        for cell in sheet[1]:
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
        
        # Apply border to all cells with data
        data_rows = sheet.max_row
        data_cols = sheet.max_column
        
        for row in range(2, data_rows + 1):
            for col in range(1, data_cols + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                cell.font = normal_font
                
                # Align numbers to the right
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column width
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    workbook.save(output_path)
